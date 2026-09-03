import json
from django.shortcuts import render
from django.views.generic import TemplateView, ListView, View
from django.http import JsonResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Count, Q
from decimal import Decimal, InvalidOperation

from inventory.models import Product, Category
from inventory.views import get_tenant_categories
from ledger.models import Customer
from pos.models import Order, OrderItem
from core_project.services import pos_checkout, add_customer_debt

class POSView(TemplateView):
    template_name = "pos.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        products = Product.objects.all().order_by('name')
        prods_list = []
        for p in products:
            prods_list.append({
                'id': p.id,
                'name': p.name,
                'sku': p.sku or '',
                'barcode': p.barcode or '',
                'category': p.category or 'عام',
                'price': str(p.selling_price),
                'stock': p.stock_quantity,
                'image': p.image.url if p.image else ''
            })
        customers = Customer.objects.all()
        custs_list = []
        for c in customers:
            custs_list.append({
                'id': c.id,
                'name': c.name,
                'phone': c.phone or '',
                'workplace': c.workplace or '',
                'balance': str(c.balance)
            })
        tenant = getattr(self.request, 'tenant', None)
        context['categories'] = get_tenant_categories(tenant) if tenant else Category.objects.all()
        context['customers'] = customers
        context['products_list_json'] = prods_list
        context['customers_list_json'] = custs_list
        return context


@method_decorator(csrf_exempt, name='dispatch')
class POSCheckoutAjaxView(View):
    """
    Handles Ajax POS orders. Receives JSON body containing cart items and checkout payment method.
    """
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            payment_method = data.get('payment_method')
            customer_id = data.get('customer_id') or None
            cart = data.get('cart', [])

            if not cart:
                return JsonResponse({'success': False, 'error': 'سلة المشتريات فارغة'}, status=400)

            import random
            import time
            order_number = f"POS-{int(time.time())}-{random.randint(10, 99)}"

            order = pos_checkout(
                order_number=order_number,
                payment_method=payment_method,
                cart_items=cart,
                customer_id=customer_id
            )

            return JsonResponse({
                'success': True,
                'order_id': order.id,
                'order_number': order.order_number,
                'total_amount': str(order.total_amount)
            })

        except ValueError as val_err:
            return JsonResponse({'success': False, 'error': str(val_err)}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f"فشلت العملية: {str(e)}"}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class AddCustomerAjaxView(View):
    """
    Quickly add a new customer directly from the POS interface.
    """
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body) if request.body else request.POST
            name = (data.get('name') or '').strip()
            phone = (data.get('phone') or '').strip() or None
            workplace = (data.get('workplace') or '').strip() or None
            address = (data.get('address') or '').strip() or None
            notes = (data.get('notes') or '').strip() or None
            initial_balance_str = (data.get('initial_balance') or '0').strip()

            if not name:
                return JsonResponse({'success': False, 'error': 'اسم العميل مطلوب ولا يمكن تركه فارغاً.'}, status=400)

            try:
                initial_balance = Decimal(str(initial_balance_str))
                if initial_balance < 0:
                    initial_balance = Decimal('0.00')
            except Exception:
                initial_balance = Decimal('0.00')

            existing = Customer.objects.filter(name=name).first()
            if existing:
                if initial_balance > 0:
                    add_customer_debt(existing.id, initial_balance, notes="إضافة رصيد مبدئي من الشاشة السريعة")
                return JsonResponse({
                    'success': True,
                    'is_existing': True,
                    'customer': {
                        'id': existing.id,
                        'name': existing.name,
                        'phone': existing.phone or '',
                        'workplace': existing.workplace or '',
                        'address': existing.address or '',
                        'balance': str(existing.balance)
                    }
                })

            customer = Customer.objects.create(
                name=name,
                phone=phone,
                workplace=workplace,
                address=address,
                notes=notes,
                balance=initial_balance
            )

            return JsonResponse({
                'success': True,
                'is_existing': False,
                'customer': {
                    'id': customer.id,
                    'name': customer.name,
                    'phone': customer.phone or '',
                    'workplace': customer.workplace or '',
                    'address': customer.address or '',
                    'balance': str(customer.balance)
                }
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': f"خطأ أثناء إضافة العميل: {str(e)}"}, status=400)


class SalesInvoicesListView(ListView):
    """
    صفحة عرض جميع فواتير المبيعات الصادرة
    """
    model = Order
    template_name = "sales_invoices.html"
    context_object_name = "orders"
    paginate_by = 50

    def get_queryset(self):
        qs = Order.objects.select_related('customer').prefetch_related('items__product').all()
        
        # Search query
        search_query = self.request.GET.get('q', '').strip()
        if search_query:
            qs = qs.filter(
                Q(order_number__icontains=search_query) |
                Q(customer__name__icontains=search_query) |
                Q(items__product__name__icontains=search_query)
            ).distinct()

        # Payment method filter
        payment_method = self.request.GET.get('payment_method', '').strip()
        if payment_method in ['cash', 'visa', 'deferred']:
            qs = qs.filter(payment_method=payment_method)

        # Date filter
        start_date = self.request.GET.get('start_date', '').strip()
        end_date = self.request.GET.get('end_date', '').strip()
        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_orders = Order.objects.all()

        total_invoices_count = all_orders.count()
        total_sales_amount = all_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00')
        cash_sales_amount = all_orders.filter(payment_method='cash').aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00')
        visa_sales_amount = all_orders.filter(payment_method='visa').aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00')
        deferred_sales_amount = all_orders.filter(payment_method='deferred').aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00')
        
        total_cogs = all_orders.aggregate(Sum('cost_of_goods_sold'))['cost_of_goods_sold__sum'] or Decimal('0.00')
        total_profit = total_sales_amount - total_cogs

        context['total_invoices_count'] = total_invoices_count
        context['total_sales_amount'] = total_sales_amount
        context['cash_visa_amount'] = cash_sales_amount + visa_sales_amount
        context['deferred_sales_amount'] = deferred_sales_amount
        context['total_profit'] = total_profit

        context['search_query'] = self.request.GET.get('q', '')
        context['selected_payment_method'] = self.request.GET.get('payment_method', '')
        context['start_date_str'] = self.request.GET.get('start_date', '')
        context['end_date_str'] = self.request.GET.get('end_date', '')

        return context


class OrderInvoiceDetailJsonView(View):
    """
    إرجاع تفاصيل الفاتورة كاملة مع أصنافها كـ JSON لمعاينتها في النافذة المنبثقة
    """
    def get(self, request, pk, *args, **kwargs):
        try:
            order = Order.objects.select_related('customer').prefetch_related('items__product').get(pk=pk)
            items_list = []
            for item in order.items.all():
                items_list.append({
                    'id': item.id,
                    'product_name': item.product.name if item.product else 'صنف محذوف',
                    'product_sku': item.product.sku if item.product else '—',
                    'quantity': item.quantity,
                    'unit_price': str(item.unit_price),
                    'total_price': str(item.unit_price * item.quantity),
                    'cost': str(item.cost),
                    'profit': str((item.unit_price - item.cost) * item.quantity),
                    'image': item.product.image.url if item.product and item.product.image else ''
                })

            customer_data = None
            if order.customer:
                customer_data = {
                    'id': order.customer.id,
                    'name': order.customer.name,
                    'phone': order.customer.phone or '',
                    'workplace': order.customer.workplace or '',
                    'address': order.customer.address or ''
                }

            profit = order.total_amount - order.cost_of_goods_sold

            return JsonResponse({
                'success': True,
                'order': {
                    'id': order.id,
                    'order_number': order.order_number,
                    'created_at': order.created_at.strftime('%Y-%m-%d %H:%M'),
                    'payment_method': order.payment_method,
                    'payment_method_display': order.get_payment_method_display(),
                    'total_amount': str(order.total_amount),
                    'cost_of_goods_sold': str(order.cost_of_goods_sold),
                    'profit': str(profit),
                    'customer': customer_data,
                    'items': items_list
                }
            })
        except Order.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'الفاتورة غير موجودة'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class PublicInvoiceDetailView(TemplateView):
    """
    صفحة عرض عامة ورسمية للفاتورة متاحة للمشاركة مع العملاء بدون الحاجة لتسجيل الدخول
    """
    template_name = "public_invoice.html"

    def get_context_data(self, **kwargs):
        from django.shortcuts import get_object_or_404
        context = super().get_context_data(**kwargs)
        order = get_object_or_404(Order.objects.select_related('customer').prefetch_related('items__product'), pk=self.kwargs['pk'])
        context['order'] = order
        return context


class ExportInvoicesExcelView(View):
    """تصدير فواتير المبيعات إلى ملف Excel أو CSV احتياطي"""
    def get(self, request, *args, **kwargs):
        from django.http import HttpResponse
        orders = Order.objects.select_related('customer').all().order_by('-created_at')

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "فواتير المبيعات"
            ws.views.sheetView[0].rightToLeft = True

            headers = ['رقم الفاتورة', 'تاريخ الإصدار', 'العميل', 'طريقة السداد', 'إجمالي الفاتورة (ج.م)', 'التكلفة (ج.م)', 'الربح الصافي (ج.م)']
            ws.append(headers)

            header_font = Font(name='Cairo', size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                                 top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

            for row_idx, o in enumerate(orders, start=2):
                profit = float(o.total_amount - o.cost_of_goods_sold)
                ws.append([
                    o.order_number,
                    o.created_at.strftime('%Y-%m-%d %H:%M'),
                    o.customer.name if o.customer else 'عميل نقدي',
                    o.get_payment_method_display(),
                    float(o.total_amount),
                    float(o.cost_of_goods_sold),
                    profit
                ])
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="right" if col_idx in [5, 6, 7] else "center", vertical="center")

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Elnamaa_Sales_Invoices.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            from django.contrib import messages
            messages.error(request, f"حدث خطأ أثناء تصدير ملف الإكسيل: {e}")
            return redirect('pos:invoices_list')


import urllib.request
import urllib.error

@method_decorator(csrf_exempt, name='dispatch')
class SyncLocalToCloudView(View):
    """
    يقوم بنقل ومزامنة الفواتير المسجلة محلياً في الكمبيوتر (Offline) إلى موقع السيرفر السحابي (PythonAnywhere)
    """
    def post(self, request, *args, **kwargs):
        try:
            cloud_sync_url = "https://webservises.pythonanywhere.com/api/v1/invoices/sync/"
            
            local_orders = Order.objects.all().prefetch_related('items__product')
            if not local_orders.exists():
                return JsonResponse({'success': True, 'synced_count': 0, 'message': 'لا توجد فواتير محلية للمزامنة'})

            invoices_payload = []
            tenant_slug = getattr(request, 'tenant', None)
            slug_str = tenant_slug.slug if tenant_slug else 'mahel'

            for order in local_orders:
                items_data = []
                for it in order.items.all():
                    items_data.append({
                        'product_id': it.product_id,
                        'quantity': it.quantity,
                        'unit_price': float(it.unit_price)
                    })
                
                if items_data:
                    invoices_payload.append({
                        'order_number': order.order_number,
                        'payment_method': order.payment_method,
                        'customer_id': order.customer_id,
                        'items': items_data
                    })

            if not invoices_payload:
                return JsonResponse({'success': True, 'synced_count': 0})

            req_body = json.dumps({
                'tenant_slug': slug_str,
                'invoices': invoices_payload
            }).encode('utf-8')

            req = urllib.request.Request(
                cloud_sync_url,
                data=req_body,
                headers={'Content-Type': 'application/json', 'User-Agent': 'AlNamaa-Desktop-Sync/1.0'}
            )

            with urllib.request.urlopen(req, timeout=15) as response:
                res_data = json.loads(response.read().decode('utf-8'))
                if res_data.get('success'):
                    return JsonResponse({
                        'success': True,
                        'synced_count': res_data.get('synced_count', len(invoices_payload))
                    })
                else:
                    return JsonResponse({'success': False, 'error': res_data.get('error', 'فشل السيرفر في معالجة المزامنة')}, status=400)

        except urllib.error.URLError as e:
            return JsonResponse({'success': False, 'error': f'تعذر الاتصال بالسيرفر السحابي (يرجى التأكد من تشغيل الإنترنت): {str(e.reason)}'}, status=503)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'خطأ أثناء المزامنة: {str(e)}'}, status=500)

