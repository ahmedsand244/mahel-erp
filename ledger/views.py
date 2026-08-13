from django.shortcuts import redirect, get_object_or_404
from django.views.generic import ListView, View, DetailView
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from .models import Customer, Supplier, Transaction
from core_project.services import (
    receive_customer_payment, send_supplier_payment,
    add_customer_debt, add_supplier_debt
)


class LedgerListView(ListView):
    model = Customer
    template_name = "ledger.html"
    context_object_name = "customers"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customers = list(Customer.objects.all())
        suppliers = list(Supplier.objects.all())

        total_customer_debt = sum((c.balance for c in customers if c.balance > 0), Decimal('0.00'))
        total_supplier_debt = sum((s.balance for s in suppliers if s.balance > 0), Decimal('0.00'))
        customers_with_debt_count = sum(1 for c in customers if c.balance > 0)
        suppliers_with_debt_count = sum(1 for s in suppliers if s.balance > 0)
        net_balance = total_customer_debt - total_supplier_debt

        context['customers'] = customers
        context['suppliers'] = suppliers
        context['total_customer_debt'] = total_customer_debt
        context['total_supplier_debt'] = total_supplier_debt
        context['net_balance'] = net_balance
        context['customers_with_debt_count'] = customers_with_debt_count
        context['suppliers_with_debt_count'] = suppliers_with_debt_count
        context['transactions'] = (
            Transaction.objects
            .select_related('customer', 'supplier')
            .order_by('-created_at')[:30]
        )
        return context


class CollectPaymentView(View):
    """تحصيل دفعة كاش من عميل — يُنقص رصيد الدين"""
    def post(self, request, *args, **kwargs):
        customer_id = request.POST.get('customer_id')
        amount = request.POST.get('amount', '').strip()
        notes = request.POST.get('notes', '').strip()

        if not customer_id or not amount:
            messages.error(request, "يرجى تحديد العميل والمبلغ بشكل صحيح.")
            return redirect('ledger:ledger_list')
        try:
            amt_dec = Decimal(str(amount))
            cust = receive_customer_payment(customer_id, amt_dec)
            tx = Transaction.objects.filter(customer=cust, transaction_type='pay_received').latest('created_at')
            if notes:
                tx.notes = notes
                tx.save(update_fields=['notes'])
            messages.success(request, f"✅ تم تحصيل {amt_dec} ج.م من العميل '{cust.name}' — الرصيد المتبقي: {cust.balance} ج.م")
        except Exception as e:
            messages.error(request, f"فشلت عملية التحصيل: {str(e)}")
        return redirect('ledger:ledger_list')


class PaySupplierView(View):
    """سداد دفعة للمورد — يُنقص رصيد ما علينا"""
    def post(self, request, *args, **kwargs):
        supplier_id = request.POST.get('supplier_id')
        amount = request.POST.get('amount', '').strip()
        notes = request.POST.get('notes', '').strip()

        if not supplier_id or not amount:
            messages.error(request, "يرجى تحديد المورد والمبلغ بشكل صحيح.")
            return redirect('ledger:ledger_list')
        try:
            amt_dec = Decimal(str(amount))
            sup = send_supplier_payment(supplier_id, amt_dec)
            tx = Transaction.objects.filter(supplier=sup, transaction_type='pay_sent').latest('created_at')
            if notes:
                tx.notes = notes
                tx.save(update_fields=['notes'])
            messages.success(request, f"✅ تم سداد {amt_dec} ج.م للمورد '{sup.name}' — الرصيد المتبقي: {sup.balance} ج.م")
        except Exception as e:
            messages.error(request, f"فشلت عملية السداد: {str(e)}")
        return redirect('ledger:ledger_list')


class AddCustomerDebtView(View):
    """إضافة دين جديد (شكك) على عميل موجود"""
    def post(self, request, pk, *args, **kwargs):
        customer = get_object_or_404(Customer, pk=pk)
        amount = request.POST.get('amount', '').strip()
        notes = request.POST.get('notes', '').strip()
        due_date = request.POST.get('due_date', '').strip() or None

        if not amount:
            messages.error(request, "يرجى إدخال مبلغ الدين.")
            return redirect('ledger:ledger_list')
        try:
            amt_dec = Decimal(str(amount))
            cust = add_customer_debt(customer.id, amt_dec, notes=notes)
            if due_date:
                cust.due_date = due_date
                cust.save(update_fields=['due_date'])
                tx = Transaction.objects.filter(customer=cust, transaction_type='sale_credit').latest('created_at')
                tx.due_date = due_date
                tx.save(update_fields=['due_date'])
            messages.success(request, f"✅ تمت إضافة دين {amt_dec} ج.م على حساب العميل '{cust.name}' — إجمالي الرصيد: {cust.balance} ج.م")
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f"خطأ: {str(e)}")
        return redirect('ledger:ledger_list')


class AddSupplierDebtView(View):
    """إضافة دين جديد على حساب مورد موجود (بضاعة آجل)"""
    def post(self, request, pk, *args, **kwargs):
        supplier = get_object_or_404(Supplier, pk=pk)
        amount = request.POST.get('amount', '').strip()
        notes = request.POST.get('notes', '').strip()
        due_date = request.POST.get('due_date', '').strip() or None

        if not amount:
            messages.error(request, "يرجى إدخال مبلغ الدين.")
            return redirect('ledger:ledger_list')
        try:
            amt_dec = Decimal(str(amount))
            sup = add_supplier_debt(supplier.id, amt_dec, notes=notes)
            if due_date:
                sup.due_date = due_date
                sup.save(update_fields=['due_date'])
                tx = Transaction.objects.filter(supplier=sup, transaction_type='purchase_credit').latest('created_at')
                tx.due_date = due_date
                tx.save(update_fields=['due_date'])
            messages.success(request, f"✅ تمت إضافة مستحقات {amt_dec} ج.م لحساب المورد '{sup.name}' — إجمالي الرصيد: {sup.balance} ج.م")
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f"خطأ: {str(e)}")
        return redirect('ledger:ledger_list')


class AddCustomerSupplierView(View):
    """
    إضافة عميل أو مورد جديد بكافة التفاصيل (الاسم، الهاتف، مكان العمل، العنوان، الصورة، الملاحظات، والرصيد المبدئي وموعد الاستحقاق).
    """
    def post(self, request, *args, **kwargs):
        party_type = request.POST.get('party_type')
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip() or None
        workplace = request.POST.get('workplace', '').strip() or None
        address = request.POST.get('address', '').strip() or None
        company = request.POST.get('company', '').strip() or None
        notes = request.POST.get('notes', '').strip() or None
        due_date = request.POST.get('due_date', '').strip() or None
        avatar = request.FILES.get('avatar')
        initial_balance_str = request.POST.get('initial_balance', '0').strip() or '0'

        if not name:
            messages.error(request, "اسم الحساب مطلوب ولا يمكن تركه فارغاً.")
            return redirect('ledger:ledger_list')

        try:
            initial_balance = Decimal(str(initial_balance_str))
            if initial_balance < 0:
                initial_balance = Decimal('0.00')
        except InvalidOperation:
            initial_balance = Decimal('0.00')

        if party_type == 'customer':
            existing = Customer.objects.filter(name=name).first()
            if existing:
                if initial_balance > 0:
                    add_customer_debt(
                        existing.id, initial_balance,
                        notes=f'إضافة دين إضافي يدوي — {initial_balance} ج.م'
                    )
                if due_date:
                    existing.due_date = due_date
                    existing.save(update_fields=['due_date'])
                messages.success(request, f"⚠️ العميل '{name}' موجود بالفعل — تمت إضافة {initial_balance} ج.م على حسابه. الإجمالي الآن: {existing.balance} ج.م")
            else:
                Customer.objects.create(
                    name=name,
                    phone=phone,
                    workplace=workplace,
                    address=address,
                    avatar=avatar,
                    notes=notes,
                    balance=initial_balance,
                    due_date=due_date
                )
                messages.success(request, f"✅ تم إنشاء حساب العميل '{name}' برصيد شكك مبدئي {initial_balance} ج.م بنجاح!")

        elif party_type == 'supplier':
            existing = Supplier.objects.filter(name=name).first()
            if existing:
                if initial_balance > 0:
                    add_supplier_debt(
                        existing.id, initial_balance,
                        notes=f'إضافة مستحقات يدوية للمورد — {initial_balance} ج.م'
                    )
                if due_date:
                    existing.due_date = due_date
                    existing.save(update_fields=['due_date'])
                messages.success(request, f"⚠️ المورد '{name}' موجود بالفعل — تمت إضافة {initial_balance} ج.م على حسابه. الإجمالي الآن: {existing.balance} ج.م")
            else:
                Supplier.objects.create(
                    name=name,
                    company=company,
                    phone=phone,
                    address=address,
                    avatar=avatar,
                    notes=notes,
                    balance=initial_balance,
                    due_date=due_date
                )
                messages.success(request, f"✅ تم إنشاء حساب المورد '{name}' برصيد مبدئي {initial_balance} ج.م بنجاح!")
        else:
            messages.error(request, "نوع الحساب غير صحيح.")

        return redirect('ledger:ledger_list')


class CustomerDetailView(DetailView):
    """صفحة تفاصيل العميل — كل معاملاته وبياناته بالتسلسل الزمني"""
    model = Customer
    template_name = 'ledger_customer_detail.html'
    context_object_name = 'customer'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['transactions'] = (
            Transaction.objects
            .filter(customer=self.object)
            .order_by('-created_at')
        )
        txs = list(context['transactions'])
        running = self.object.balance
        for tx in txs:
            tx.running_balance = running
            if tx.transaction_type == 'pay_received':
                running += tx.amount
            else:
                running -= tx.amount
        context['transactions'] = txs
        return context


class SupplierDetailView(DetailView):
    """صفحة تفاصيل المورد — كل معاملاته وبياناته بالتسلسل الزمني"""
    model = Supplier
    template_name = 'ledger_supplier_detail.html'
    context_object_name = 'supplier'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['transactions'] = (
            Transaction.objects
            .filter(supplier=self.object)
            .order_by('-created_at')
        )
        txs = list(context['transactions'])
        running = self.object.balance
        for tx in txs:
            tx.running_balance = running
            if tx.transaction_type == 'pay_sent':
                running += tx.amount
            else:
                running -= tx.amount
        context['transactions'] = txs
        return context


class UpdateCustomerProfileView(View):
    """تحديث الملف الشخصي وبيانات العميل التفصيلية"""
    def post(self, request, pk, *args, **kwargs):
        customer = get_object_or_404(Customer, pk=pk)
        customer.name = request.POST.get('name', customer.name).strip()
        customer.phone = request.POST.get('phone', '').strip() or None
        customer.workplace = request.POST.get('workplace', '').strip() or None
        customer.address = request.POST.get('address', '').strip() or None
        customer.notes = request.POST.get('notes', '').strip() or None
        
        due_date_str = request.POST.get('due_date', '').strip()
        customer.due_date = due_date_str if due_date_str else None
        
        if 'avatar' in request.FILES:
            customer.avatar = request.FILES['avatar']
            
        customer.save()
        messages.success(request, f"✅ تم تحديث ملف العميل '{customer.name}' بنجاح!")
        return redirect('ledger:customer_detail', pk=pk)


class UpdateSupplierProfileView(View):
    """تحديث الملف الشخصي وبيانات المورد التفصيلية"""
    def post(self, request, pk, *args, **kwargs):
        supplier = get_object_or_404(Supplier, pk=pk)
        supplier.name = request.POST.get('name', supplier.name).strip()
        supplier.company = request.POST.get('company', '').strip() or None
        supplier.phone = request.POST.get('phone', '').strip() or None
        supplier.address = request.POST.get('address', '').strip() or None
        supplier.notes = request.POST.get('notes', '').strip() or None
        
        due_date_str = request.POST.get('due_date', '').strip()
        supplier.due_date = due_date_str if due_date_str else None
        
        if 'avatar' in request.FILES:
            supplier.avatar = request.FILES['avatar']
            
        supplier.save()
        messages.success(request, f"✅ تم تحديث ملف المورد '{supplier.name}' بنجاح!")
        return redirect('ledger:supplier_detail', pk=pk)


class PublicCustomerLedgerView(DetailView):
    """صفحة كشف حساب عامة رسمية للعميل صادرة ومتاحة للمشاركة بدون تسجيل دخول"""
    model = Customer
    template_name = 'public_customer_ledger.html'
    context_object_name = 'customer'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['transactions'] = (
            Transaction.objects
            .filter(customer=self.object)
            .order_by('-created_at')
        )
        return context


class ExportCustomersExcelView(View):
    """تصدير سجلات كشف حساب جميع العملاء والديون إلى ملف Excel"""
    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ديون العملاء"
        ws.views.sheetView[0].rightToLeft = True

        headers = ['اسم العميل', 'رقم الهاتف', 'مكان العمل / المزرعة', 'العنوان', 'رصيد الدين المستحق (ج.م)', 'موعد الاستحقاق']
        ws.append(headers)

        header_font = Font(name='Cairo', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        customers = Customer.objects.all().order_by('-balance')
        thin_border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                             top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

        for row_idx, c in enumerate(customers, start=2):
            ws.append([
                c.name,
                c.phone or '—',
                c.workplace or '—',
                c.address or '—',
                float(c.balance),
                str(c.due_date) if c.due_date else '—'
            ])
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right" if col_idx == 1 else "center", vertical="center")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Elnamaa_Customers_Debts.xlsx"'
        wb.save(response)
        return response


class ExportSuppliersExcelView(View):
    """تصدير سجلات كشف حساب جميع الموردين والمستحقات إلى ملف Excel"""
    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "مستحقات الموردين"
        ws.views.sheetView[0].rightToLeft = True

        headers = ['اسم المورد', 'الشركة / المصنع', 'رقم الهاتف', 'العنوان', 'المستحقات له علينا (ج.م)', 'موعد الاستحقاق']
        ws.append(headers)

        header_font = Font(name='Cairo', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        suppliers = Supplier.objects.all().order_by('-balance')
        thin_border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                             top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

        for row_idx, s in enumerate(suppliers, start=2):
            ws.append([
                s.name,
                s.company or '—',
                s.phone or '—',
                s.address or '—',
                float(s.balance),
                str(s.due_date) if s.due_date else '—'
            ])
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right" if col_idx == 1 else "center", vertical="center")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Elnamaa_Suppliers_Debts.xlsx"'
        wb.save(response)
        return response

