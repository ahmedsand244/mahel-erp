import json
import uuid
from django.shortcuts import redirect, get_object_or_404, render
from django.views.generic import ListView, CreateView, UpdateView, DetailView, View, TemplateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import F
from django.db import transaction
from .models import Product, StockAlert, PurchaseOrder, PurchaseOrderItem, Category
from ledger.models import Supplier


def get_tenant_categories(tenant):
    """جلب فئات المنتجات للشركة وإعادة إنشاء فئات افتراضية أولية إذا كانت فارغة."""
    categories = Category.objects.filter(tenant=tenant).order_by('name')
    if not categories.exists():
        default_names = ['أسمدة ومخصبات زراعية', 'مبيدات حشرية وفطرية', 'قطع غيار مواقير ورش', 'معدات وآلات زراعية', 'زيوت وشحومات', 'عام / متنوع']
        for dname in default_names:
            Category.objects.create(name=dname, tenant=tenant)
        categories = Category.objects.filter(tenant=tenant).order_by('name')
    return categories


class InventoryListView(ListView):
    model = Product
    template_name = "inventory.html"
    context_object_name = "products"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Dynamic, real-time deduplicated stock alerts
        low_stock_products = Product.objects.filter(
            stock_quantity__lte=F('min_stock_threshold')
        ).order_by('stock_quantity', 'name')
        
        tenant = getattr(self.request, 'tenant', None)
        context['categories'] = get_tenant_categories(tenant) if tenant else Category.objects.all()
        context['low_stock_products'] = low_stock_products
        context['out_of_stock_count'] = low_stock_products.filter(stock_quantity__lte=0).count()
        context['low_stock_count'] = low_stock_products.filter(stock_quantity__gt=0).count()
        context['category_choices'] = Product.CATEGORY_CHOICES
        return context


class ProductCreateView(CreateView):
    model = Product
    fields = ['name', 'sku', 'barcode', 'category', 'image', 'purchase_price', 'selling_price', 'stock_quantity', 'min_stock_threshold']
    template_name = "inventory.html"
    success_url = reverse_lazy('inventory:inventory_list')

    def form_valid(self, form):
        if form.cleaned_data.get('barcode') == '':
            form.instance.barcode = None
        messages.success(self.request, f"تمت إضافة المنتج '{form.instance.name}' بنجاح إلى المخزن!")
        return super().form_valid(form)

    def form_invalid(self, form):
        error_msg = "; ".join([f"{', '.join(errs)}" for field, errs in form.errors.items()])
        messages.error(self.request, f"خطأ أثناء إضافة المنتج: {error_msg}")
        return redirect('inventory:inventory_list')


class ProductUpdateView(UpdateView):
    model = Product
    fields = ['name', 'sku', 'barcode', 'category', 'image', 'purchase_price', 'selling_price', 'stock_quantity', 'min_stock_threshold']
    template_name = "inventory.html"
    success_url = reverse_lazy('inventory:inventory_list')

    def form_valid(self, form):
        if form.cleaned_data.get('barcode') == '':
            form.instance.barcode = None
        messages.success(self.request, f"تم تحديث بيانات المنتج '{form.instance.name}' بنجاح!")
        return super().form_valid(form)

    def form_invalid(self, form):
        error_msg = "; ".join([f"{', '.join(errs)}" for field, errs in form.errors.items()])
        messages.error(self.request, f"خطأ أثناء تعديل المنتج: {error_msg}")
        return redirect('inventory:inventory_list')


class ProductDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        product = get_object_or_404(Product, pk=pk)
        name = product.name
        try:
            product.delete()
            messages.success(request, f"تم حذف المنتج '{name}' من المخزن بنجاح.")
        except Exception:
            messages.error(request, f"لا يمكن حذف المنتج '{name}' لأنه مرتبط بفواتير أو عمليات صيانة مسجلة.")
        return redirect('inventory:inventory_list')


class QuickRestockView(View):
    """
    Quickly add stock quantity to a product directly from Stock Alerts or table.
    """
    def post(self, request, pk, *args, **kwargs):
        product = get_object_or_404(Product, pk=pk)
        added_qty_str = request.POST.get('added_quantity', '0').strip()

        try:
            added_qty = int(added_qty_str)
            if added_qty <= 0:
                messages.error(request, "يرجى إدخال كمية صحيحة أكبر من صفر.")
                return redirect('inventory:inventory_list')
        except ValueError:
            messages.error(request, "كمية غير صحيحة.")
            return redirect('inventory:inventory_list')

        product.stock_quantity += added_qty
        product.save()

        # Mark alerts resolved for this product if any
        StockAlert.objects.filter(product=product, is_resolved=False).update(is_resolved=True)

        messages.success(
            request, 
            f"📦 تم تزويد مخزون '{product.name}' بـ (+{added_qty} قطع) بنجاح! الرصيد الحالي المتوفر الآن: {product.stock_quantity} قطعة."
        )
        return redirect('inventory:inventory_list')


class BulkPriceAdjustmentView(View):
    """
    POST view to adjust prices by a percentage rate (e.g. +5%, +10%, +15%, -5%).
    """
    def post(self, request, *args, **kwargs):
        percentage_str = request.POST.get('percentage', '0').strip()
        price_target = request.POST.get('price_target', 'selling')
        scope = request.POST.get('scope', 'all')
        category = request.POST.get('category', '')
        selected_ids = request.POST.getlist('selected_products')

        try:
            pct = Decimal(percentage_str)
        except Exception:
            messages.error(request, "يرجى إدخال نسبة مئوية صحيحة (مثال: 5 أو 10 أو 12.5).")
            return redirect('inventory:inventory_list')

        if pct == 0:
            messages.warning(request, "النسبة المئوية 0% — لم يتم تغيير أي أسعار.")
            return redirect('inventory:inventory_list')

        queryset = Product.objects.all()
        if scope == 'selected':
            if not selected_ids:
                messages.error(request, "لم تقم بتحديد أي منتجات لتطبيق الزيادة عليها. علم على المنتجات المطلوبة أولاً.")
                return redirect('inventory:inventory_list')
            queryset = queryset.filter(id__in=selected_ids)
        elif scope == 'category' and category:
            queryset = queryset.filter(category=category)

        multiplier = Decimal('1.00') + (pct / Decimal('100.00'))

        updated_count = 0
        for product in queryset:
            if price_target in ['selling', 'both']:
                new_sell = (product.selling_price * multiplier).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                product.selling_price = max(Decimal('0.00'), new_sell)
            if price_target in ['purchase', 'both']:
                new_buy = (product.purchase_price * multiplier).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                product.purchase_price = max(Decimal('0.00'), new_buy)
            product.save()
            updated_count += 1

        sign = "+" if pct > 0 else ""
        target_desc = "سعر البيع" if price_target == 'selling' else ("سعر الشراء" if price_target == 'purchase' else "سعر البيع والشراء")
        messages.success(
            request,
            f"🚀 تم تحديث {target_desc} لـ {updated_count} منتج بنسبة ({sign}{pct}%) بنجاح!"
        )
        return redirect('inventory:inventory_list')


class SingleProductPriceAdjustmentView(View):
    """
    Quick percentage adjustment for a single specific product.
    """
    def post(self, request, pk, *args, **kwargs):
        product = get_object_or_404(Product, pk=pk)
        percentage_str = request.POST.get('percentage', '0').strip()
        price_target = request.POST.get('price_target', 'selling')

        try:
            pct = Decimal(percentage_str)
        except Exception:
            messages.error(request, "يرجى إدخال نسبة صحيحة.")
            return redirect('inventory:inventory_list')

        multiplier = Decimal('1.00') + (pct / Decimal('100.00'))

        if price_target in ['selling', 'both']:
            product.selling_price = max(Decimal('0.00'), (product.selling_price * multiplier).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        if price_target in ['purchase', 'both']:
            product.purchase_price = max(Decimal('0.00'), (product.purchase_price * multiplier).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        
        product.save()
        sign = "+" if pct > 0 else ""
        messages.success(request, f"تم تعديل سعر المنتج '{product.name}' بنسبة ({sign}{pct}%) — سعر البيع الجديد: {product.selling_price} ج.م")
        return redirect('inventory:inventory_list')


class PurchaseOrderListView(ListView):
    model = PurchaseOrder
    template_name = "purchase_order_list.html"
    context_object_name = "orders"

    def get_queryset(self):
        qs = super().get_queryset().select_related('supplier').prefetch_related('items', 'items__product', 'items__supplier')
        status_filter = self.request.GET.get('status')
        supplier_filter = self.request.GET.get('supplier_id')
        if status_filter:
            qs = qs.filter(status=status_filter)
        if supplier_filter and str(supplier_filter).isdigit():
            qs = qs.filter(supplier_id=int(supplier_filter))
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['suppliers'] = Supplier.objects.all()
        context['draft_count'] = PurchaseOrder.objects.filter(status='draft').count()
        context['sent_count'] = PurchaseOrder.objects.filter(status='sent').count()
        context['received_count'] = PurchaseOrder.objects.filter(status='received').count()
        context['critical_stock_count'] = Product.objects.filter(stock_quantity__lte=F('min_stock_threshold')).count()
        return context


class PurchaseOrderBuilderView(View):
    def get(self, request, pk=None, *args, **kwargs):
        order = None
        if pk:
            order = get_object_or_404(PurchaseOrder, pk=pk)

        all_products = Product.objects.all().select_related('default_supplier')
        low_stock_products = Product.objects.filter(
            stock_quantity__lte=F('min_stock_threshold')
        ).select_related('default_supplier').order_by('stock_quantity', 'name')
        suppliers = Supplier.objects.all()

        context = {
            'order': order,
            'all_products': all_products,
            'low_stock_products': low_stock_products,
            'suppliers': suppliers,
        }
        return render(request, "purchase_order_builder.html", context)

    def post(self, request, pk=None, *args, **kwargs):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'بيانات غير صحيحة'}, status=400)

        supplier_id = data.get('supplier_id') or None
        notes = data.get('notes', '')
        status = data.get('status', 'draft')
        items_data = data.get('items', [])

        if not items_data:
            return JsonResponse({'success': False, 'error': 'يرجى إضافة صنف واحد على الأقل للطلبية'}, status=400)

        with transaction.atomic():
            if pk:
                order = get_object_or_404(PurchaseOrder, pk=pk)
                order.supplier_id = supplier_id
                order.notes = notes
                order.status = status
                order.save()
                order.items.all().delete()
            else:
                today_str = timezone.now().strftime('%Y%m%d')
                random_suffix = str(uuid.uuid4().hex[:4]).upper()
                order_number = f"PO-{today_str}-{random_suffix}"
                order = PurchaseOrder.objects.create(
                    order_number=order_number,
                    supplier_id=supplier_id,
                    notes=notes,
                    status=status,
                )

            for item in items_data:
                product_id = item.get('product_id')
                custom_name = str(item.get('custom_name', '') or '').strip()
                item_supplier_id = item.get('supplier_id')
                quantity = int(item.get('quantity', 1))
                unit_cost = Decimal(str(item.get('unit_cost', 0)))

                prod_obj = Product.objects.filter(id=product_id).first() if product_id else None

                PurchaseOrderItem.objects.create(
                    purchase_order=order,
                    product=prod_obj,
                    custom_item_name=custom_name or (prod_obj.name if prod_obj else "صنف مخصص"),
                    supplier_id=item_supplier_id if item_supplier_id else None,
                    quantity_requested=max(1, quantity),
                    unit_cost=max(Decimal('0.00'), unit_cost)
                )

            if status == 'received' and not order.received_at:
                for item in order.items.all():
                    if item.product and not item.is_received:
                        item.product.stock_quantity += item.quantity_requested
                        item.product.save()
                        StockAlert.objects.filter(product=item.product, is_resolved=False).update(is_resolved=True)
                        item.is_received = True
                        item.save()
                order.received_at = timezone.now()
                order.save()

        return JsonResponse({
            'success': True,
            'order_id': order.id,
            'order_number': order.order_number,
            'redirect_url': str(reverse_lazy('inventory:purchase_order_detail', kwargs={'pk': order.id}))
        })


class PurchaseOrderDetailView(DetailView):
    model = PurchaseOrder
    template_name = "purchase_order_detail.html"
    context_object_name = "order"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        order = self.object
        items = order.items.select_related('product', 'supplier', 'product__default_supplier').all()

        grouped_by_supplier = {}
        for item in items:
            supp_name = item.display_supplier_name
            supp_phone = ""
            if item.supplier and item.supplier.phone:
                supp_phone = item.supplier.phone
            elif item.product and item.product.default_supplier and item.product.default_supplier.phone:
                supp_phone = item.product.default_supplier.phone
            elif order.supplier and order.supplier.phone:
                supp_phone = order.supplier.phone

            if supp_name not in grouped_by_supplier:
                grouped_by_supplier[supp_name] = {
                    'supplier_name': supp_name,
                    'phone': supp_phone,
                    'items': [],
                    'total_cost': Decimal('0.00'),
                }
            grouped_by_supplier[supp_name]['items'].append(item)
            grouped_by_supplier[supp_name]['total_cost'] += item.total_cost

        context['grouped_items'] = grouped_by_supplier
        return context


class PurchaseOrderReceiveView(View):
    def post(self, request, pk, *args, **kwargs):
        order = get_object_or_404(PurchaseOrder, pk=pk)

        if order.status == 'received':
            messages.warning(request, f"الطلبية رقم {order.order_number} تم استلامها وإضافتها للمخزن سابقاً!")
            return redirect('inventory:purchase_order_detail', pk=order.id)

        with transaction.atomic():
            updated_count = 0
            for item in order.items.all():
                if item.product and not item.is_received:
                    item.product.stock_quantity += item.quantity_requested
                    item.product.save()
                    StockAlert.objects.filter(product=item.product, is_resolved=False).update(is_resolved=True)
                    item.is_received = True
                    item.save()
                    updated_count += 1

            order.status = 'received'
            order.received_at = timezone.now()
            order.save()

        messages.success(
            request, 
            f"🎉 تم استلام شحنة الطلبية ({order.order_number}) بنجاح! تم تزويد رصيد المخزون لـ ({updated_count}) صنف."
        )
        return redirect('inventory:purchase_order_detail', pk=order.id)


class PurchaseOrderDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        num = order.order_number
        order.delete()
        messages.success(request, f"تم حذف طلب البضاعة '{num}' بنجاح.")
        return redirect('inventory:purchase_order_list')


class ExportProductsExcelView(View):
    """تصدير جميع منتجات المخزن إلى ملف Excel (.xlsx)"""
    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "منتجات المخزن"
        ws.views.sheetView[0].rightToLeft = True

        # Headers
        headers = ['اسم المنتج', 'الباركوود', 'التصنيف', 'سعر الشراء (ج.م)', 'سعر البيع (ج.م)', 'الكمية بالمخزن', 'الحد الأدنى']
        ws.append(headers)

        # Style headers
        header_font = Font(name='Cairo', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                             top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        products = Product.objects.all().order_by('name')
        for row_idx, p in enumerate(products, start=2):
            ws.append([
                p.name,
                p.barcode or p.sku or '',
                p.get_category_display() if hasattr(p, 'get_category_display') else p.category,
                float(p.purchase_price),
                float(p.selling_price),
                p.stock_quantity,
                p.min_stock_threshold
            ])
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx in [4, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right" if col_idx == 1 else "center", vertical="center")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Elnamaa_Inventory_Products.xlsx"'
        wb.save(response)
        return response


class DownloadSampleProductsExcelView(View):
    """تحميل نموذج استيراد المنتجات Excel التوضيحي"""
    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "نموذج استيراد المنتجات"
        ws.views.sheetView[0].rightToLeft = True

        headers = ['اسم المنتج', 'الباركوود', 'التصنيف', 'سعر الشراء (ج.م)', 'سعر البيع (ج.م)', 'الكمية بالمخزن', 'الحد الأدنى']
        ws.append(headers)

        header_font = Font(name='Cairo', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sample_rows = [
            ['موتور مياه 1 حصان إيطالي', '62910001', 'مواتير', 3500, 4200, 10, 2],
            ['زيت شل 16 لتر هيدروليك', '62910002', 'زيوت وفلاتر', 1800, 2100, 15, 3],
            ['سير كبولة متعدد المقاسات', '62910003', 'سيور وسلاسل', 250, 320, 30, 5],
        ]
        for row in sample_rows:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Elnamaa_Products_Import_Sample.xlsx"'
        wb.save(response)
        return response


class ImportProductsExcelView(View):
    """استيراد ملف Excel أو CSV وتحديث/إضافة المنتجات دفعة واحدة"""
    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "يرجى اختيار ملف Excel لاستيراد المنتجات.")
            return redirect('inventory:inventory_list')

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active

            imported_count = 0
            updated_count = 0

            with transaction.atomic():
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) <= 1:
                    messages.warning(request, "الملف المرفوع فارغ أو لا يحتوي على صفوف بيانات!")
                    return redirect('inventory:inventory_list')

                # Skip header row (row 0)
                for row in rows[1:]:
                    if not row or not row[0]:
                        continue

                    name = str(row[0]).strip()
                    barcode = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
                    if barcode in ['None', '', '-']:
                        barcode = None

                    category = str(row[2]).strip() if len(row) > 2 and row[2] else 'قطع غيار عامة'
                    
                    try:
                        purchase_price = Decimal(str(row[3])) if len(row) > 3 and row[3] is not None else Decimal('0.00')
                    except Exception:
                        purchase_price = Decimal('0.00')

                    try:
                        selling_price = Decimal(str(row[4])) if len(row) > 4 and row[4] is not None else Decimal('0.00')
                    except Exception:
                        selling_price = Decimal('0.00')

                    try:
                        stock_qty = int(float(row[5])) if len(row) > 5 and row[5] is not None else 0
                    except Exception:
                        stock_qty = 0

                    try:
                        min_stock = int(float(row[6])) if len(row) > 6 and row[6] is not None else 5
                    except Exception:
                        min_stock = 5

                    prod, created = Product.objects.get_or_create(
                        name=name,
                        defaults={
                            'barcode': barcode,
                            'category': category,
                            'purchase_price': purchase_price,
                            'selling_price': selling_price,
                            'stock_quantity': stock_qty,
                            'min_stock_threshold': min_stock,
                        }
                    )

                    if not created:
                        if barcode: prod.barcode = barcode
                        prod.category = category
                        prod.purchase_price = purchase_price
                        prod.selling_price = selling_price
                        prod.stock_quantity = stock_qty
                        prod.min_stock_threshold = min_stock
                        prod.save()
                        updated_count += 1
                    else:
                        imported_count += 1

            messages.success(
                request,
                f"🎉 تم بنجاح استيراد ({imported_count}) منتج جديد، وتحديث بيانات ({updated_count}) منتج سابق في المخزن!"
            )
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء قراءة ملف Excel: {str(e)}")

        return redirect('inventory:inventory_list')


class BarcodeGeneratorView(TemplateView):
    """مولد وطباعة ملصقات الباركود للمنتجات وأجهزة الورشة"""
    template_name = "barcode_labels.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        products = Product.objects.all().order_by('name')
        
        # Format JSON products for instant Alpine JS selector
        prod_json = []
        for p in products:
            prod_json.append({
                'id': p.id,
                'name': p.name,
                'barcode': p.barcode or p.sku or f"PROD-{p.id}",
                'price': str(p.selling_price),
                'category': str(p.category) if p.category else 'عام'
            })
            
        context['products'] = products
        context['products_json'] = prod_json
        
        selected_id = self.request.GET.get('product_id')
        selected_prod = None
        if selected_id and str(selected_id).isdigit():
            selected_prod = products.filter(id=int(selected_id)).first()

        context['selected_product'] = selected_prod
        context['initial_id'] = selected_prod.id if selected_prod else ''
        context['initial_title'] = selected_prod.name if selected_prod else ''
        context['initial_barcode'] = (selected_prod.barcode or selected_prod.sku or f"PROD-{selected_prod.id}") if selected_prod else '62910001'
        context['initial_price'] = str(selected_prod.selling_price) if selected_prod else '0.00'
        return context


class CategoryCreateView(View):
    def post(self, request):
        name = request.POST.get('name', '').strip()
        if name:
            tenant = getattr(request, 'tenant', None)
            category, created = Category.objects.get_or_create(name=name, tenant=tenant)
            if created:
                messages.success(request, f'🎉 تم إضافة الفئة "{name}" بنجاح.')
            else:
                messages.info(request, f'الفئة "{name}" موجودة بالفعل.')
        else:
            messages.error(request, 'الرجاء إدخال اسم الفئة.')
        return redirect(request.META.get('HTTP_REFERER', 'inventory:inventory_list'))


class CategoryDeleteView(View):
    def post(self, request, pk):
        category = get_object_or_404(Category, pk=pk)
        name = category.name
        category.delete()
        messages.success(request, f'🗑️ تم حذف الفئة "{name}" بنجاح.')
        return redirect(request.META.get('HTTP_REFERER', 'inventory:inventory_list'))

